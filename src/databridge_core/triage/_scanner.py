"""Batch Excel scanner — fast metadata extraction via openpyxl.

Scans Excel files for structural metadata (sheet counts, formula counts,
named ranges, macros) without full formula decomposition. Uses
read_only=True for streaming speed and ThreadPoolExecutor for concurrency.
"""
from __future__ import annotations

import logging
import os
import re
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple

from ._types import (
    Archetype,
    FileTriageResult,
    ScanStatus,
    SheetMetadata,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Formula regex patterns (compiled once)
# ---------------------------------------------------------------------------
_SUMIF_RE = re.compile(r"\bSUMIFS?\b", re.IGNORECASE)
_VLOOKUP_RE = re.compile(r"\bVLOOKUP\b", re.IGNORECASE)
_IF_CHAIN_RE = re.compile(r"\bIF\b.*\bIF\b", re.IGNORECASE | re.DOTALL)
_FUNC_RE = re.compile(r"\b([A-Z]{2,})\s*\(", re.IGNORECASE)

# Extensions openpyxl can handle
_SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
_SKIP_EXTENSIONS = {".xlsb", ".xls"}

# Sampling limits for speed
_MAX_SAMPLE_ROWS = 200
_MAX_SAMPLE_COLS = 50
_MAX_ANCHOR_ROWS = 20
_MAX_ANCHOR_COLS = 20


class BatchExcelScanner:
    """Scan Excel files for triage metadata.

    Parameters
    ----------
    max_workers : int
        Number of threads for concurrent scanning (default 4).
    deep_scan : bool
        If True, deep scan fields are populated (no-op in databridge-core;
        BLCE ExcelLogicExtractor is only available in the full DataBridge).
    """

    def __init__(self, max_workers: int = 4, deep_scan: bool = False) -> None:
        self.max_workers = max_workers
        self.deep_scan = deep_scan

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def scan_directory(
        self,
        directory: str,
        progress_callback: Optional[Callable[[int, int, str], None]] = None,
    ) -> List[FileTriageResult]:
        """Scan all Excel files in *directory* (non-recursive by default).

        Parameters
        ----------
        directory : str
            Path to the directory containing Excel files.
        progress_callback : callable, optional
            Called as ``callback(completed, total, file_name)`` after each file.

        Returns
        -------
        list[FileTriageResult]
        """
        dir_path = Path(directory)
        if not dir_path.is_dir():
            raise FileNotFoundError(f"Directory not found: {directory}")

        files = self._collect_files(dir_path)
        total = len(files)
        results: List[FileTriageResult] = []

        if total == 0:
            return results

        with ThreadPoolExecutor(max_workers=self.max_workers) as pool:
            futures = {pool.submit(self.scan_file, str(f)): f for f in files}
            for idx, future in enumerate(as_completed(futures), 1):
                file_path = futures[future]
                try:
                    result = future.result()
                except Exception as exc:
                    result = FileTriageResult(
                        file_path=str(file_path),
                        file_name=file_path.name,
                        file_extension=file_path.suffix.lower(),
                        scan_status=ScanStatus.ERROR,
                        error_message=f"Unexpected thread error: {exc}",
                    )
                results.append(result)
                if progress_callback:
                    progress_callback(idx, total, file_path.name)

        return results

    def scan_file(self, file_path: str) -> FileTriageResult:
        """Scan a single Excel file and return its triage metadata."""
        path = Path(file_path)
        ext = path.suffix.lower()

        # Basic file info
        try:
            stat = os.stat(file_path)
            file_size = stat.st_size
        except OSError as exc:
            return FileTriageResult(
                file_path=str(path),
                file_name=path.name,
                file_extension=ext,
                scan_status=ScanStatus.ERROR,
                error_message=f"Cannot stat file: {exc}",
            )

        # Skip unsupported formats
        if ext in _SKIP_EXTENSIONS:
            return FileTriageResult(
                file_path=str(path),
                file_name=path.name,
                file_size_bytes=file_size,
                file_extension=ext,
                scan_status=ScanStatus.SKIPPED,
                error_message=f"Unsupported format: {ext} (openpyxl cannot read {ext} files)",
            )

        if ext not in _SUPPORTED_EXTENSIONS:
            return FileTriageResult(
                file_path=str(path),
                file_name=path.name,
                file_size_bytes=file_size,
                file_extension=ext,
                scan_status=ScanStatus.SKIPPED,
                error_message=f"Not an Excel file: {ext}",
            )

        # Open workbook
        try:
            import openpyxl
        except ImportError:
            return FileTriageResult(
                file_path=str(path),
                file_name=path.name,
                file_size_bytes=file_size,
                file_extension=ext,
                scan_status=ScanStatus.ERROR,
                error_message="openpyxl not installed",
            )

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=False)
        except Exception as exc:
            exc_type = type(exc).__name__
            return FileTriageResult(
                file_path=str(path),
                file_name=path.name,
                file_size_bytes=file_size,
                file_extension=ext,
                scan_status=ScanStatus.ERROR,
                error_message=f"{exc_type}: {exc}",
            )

        try:
            return self._extract_metadata(wb, path, file_size, ext)
        except Exception as exc:
            return FileTriageResult(
                file_path=str(path),
                file_name=path.name,
                file_size_bytes=file_size,
                file_extension=ext,
                scan_status=ScanStatus.ERROR,
                error_message=f"Extraction error: {exc}",
            )
        finally:
            try:
                wb.close()
            except Exception:
                pass

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _collect_files(self, directory: Path) -> List[Path]:
        """Collect all Excel-like files in directory (non-recursive)."""
        all_ext = _SUPPORTED_EXTENSIONS | _SKIP_EXTENSIONS
        files = []
        for item in directory.iterdir():
            if item.is_file() and item.suffix.lower() in all_ext:
                files.append(item)
        return sorted(files, key=lambda p: p.name.lower())

    def _extract_metadata(
        self,
        wb: Any,
        path: Path,
        file_size: int,
        ext: str,
    ) -> FileTriageResult:
        """Extract metadata from an open openpyxl Workbook (read_only mode)."""
        sheet_names = wb.sheetnames
        sheets: List[SheetMetadata] = []
        total_formula_count = 0
        total_row_count = 0
        has_sumif = False
        has_vlookup = False
        has_if_chain = False
        func_counter: Counter = Counter()

        for ws in wb.worksheets:
            sheet_meta, formulas = self._scan_sheet(ws)
            sheets.append(sheet_meta)
            total_formula_count += sheet_meta.formula_count
            total_row_count += sheet_meta.row_count

            # Analyze formula patterns from sampled formulas
            for formula in formulas:
                if _SUMIF_RE.search(formula):
                    has_sumif = True
                if _VLOOKUP_RE.search(formula):
                    has_vlookup = True
                if _IF_CHAIN_RE.search(formula):
                    has_if_chain = True
                for match in _FUNC_RE.finditer(formula):
                    func_counter[match.group(1).upper()] += 1

        # Named ranges
        try:
            named_range_count = len(wb.defined_names)
        except Exception:
            named_range_count = 0

        # Macros (vba_archive available in read_only mode)
        has_macros = getattr(wb, "vba_archive", None) is not None

        # Pivot tables
        has_pivot_tables = any(s.has_pivot_table for s in sheets)

        # Top formula functions
        dominant_functions = [f for f, _ in func_counter.most_common(5)]

        # Deep scan (no-op in databridge-core — BLCE not available)
        measure_count = None
        dependency_count = None
        confidence = None
        if self.deep_scan:
            measure_count, dependency_count, confidence = self._run_deep_scan(str(path))

        return FileTriageResult(
            file_path=str(path),
            file_name=path.name,
            file_size_bytes=file_size,
            file_extension=ext,
            scan_status=ScanStatus.OK,
            sheet_count=len(sheet_names),
            sheet_names=sheet_names,
            sheets=sheets,
            total_row_count=total_row_count,
            formula_count=total_formula_count,
            named_range_count=named_range_count,
            has_macros=has_macros,
            has_pivot_tables=has_pivot_tables,
            measure_count=measure_count,
            dependency_count=dependency_count,
            confidence=confidence,
            has_sumif_pattern=has_sumif,
            has_vlookup_pattern=has_vlookup,
            has_if_chain=has_if_chain,
            dominant_formula_functions=dominant_functions,
        )

    def _scan_sheet(self, ws: Any) -> Tuple[SheetMetadata, List[str]]:
        """Extract metadata from a single worksheet (read_only compatible).

        Returns (SheetMetadata, list_of_formula_strings) so the caller can
        analyze formula patterns without re-iterating.
        """
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        # Detect empty sheet
        if max_row == 0 and max_col == 0:
            return SheetMetadata(name=ws.title, is_empty=True), []

        # Single-pass scan: anchor detection + formula collection
        formula_count = 0
        anchor_row = None
        anchor_col = None
        formulas: List[str] = []

        sample_rows = min(max_row, _MAX_SAMPLE_ROWS)
        sample_cols = min(max_col, _MAX_SAMPLE_COLS)

        row_idx = 0
        for row in ws.iter_rows(
            min_row=1, max_row=sample_rows,
            min_col=1, max_col=sample_cols,
        ):
            row_idx += 1
            non_empty = 0
            for cell in row:
                val = cell.value
                if val is not None:
                    non_empty += 1
                if val and isinstance(val, str) and val.startswith("="):
                    formula_count += 1
                    formulas.append(val)

            # Anchor: first row where >50% of sampled cols are non-empty
            if anchor_row is None and row_idx <= _MAX_ANCHOR_ROWS:
                check_cols = min(sample_cols, _MAX_ANCHOR_COLS)
                if check_cols > 0 and non_empty / check_cols > 0.5:
                    anchor_row = row_idx
                    anchor_col = 1

        # Extrapolate formula count if we sampled
        if max_row > sample_rows and formula_count > 0:
            formula_count = int(formula_count * (max_row / sample_rows))

        # Pivot table detection (not available in read_only, check safely)
        has_pivot = bool(getattr(ws, "_pivots", None))

        return SheetMetadata(
            name=ws.title,
            row_count=max_row,
            col_count=max_col,
            formula_count=formula_count,
            anchor_row=anchor_row,
            anchor_col=anchor_col,
            has_pivot_table=has_pivot,
            is_empty=(max_row <= 1 and formula_count == 0),
        ), formulas

    def _run_deep_scan(self, file_path: str) -> Tuple[int, int, float]:
        """Deep scan stub — BLCE ExcelLogicExtractor is not available in databridge-core.

        Returns zeros. For full deep-scan capability, use the main DataBridge package.
        """
        logger.info(
            "Deep scan requested for %s but BLCE is not available in databridge-core. "
            "Returning zeros.",
            file_path,
        )
        return (0, 0, 0.0)
